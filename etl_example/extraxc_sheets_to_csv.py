import openpyxl
import pandas as pd
import numpy as np
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

original_files = {
    "City1": os.path.join(BASE_DIR, "data", "City1", "raw_data", "Source File City1.xlsx"),
    "City2": os.path.join(BASE_DIR, "data", "City2", "raw_data", "Source File City2.xlsx"),
}

price_col_name = "Revenue"
col_name_info = "Source"

def process_city(city: str, xlsx_path: str):
    if not os.path.exists(xlsx_path):
        print(f"File not found for {city}: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames

    output_folder = os.path.join(BASE_DIR, "data", city, "extracted_data")
    os.makedirs(output_folder, exist_ok=True)

    print(f"\nProcessing {city.upper()} ({len(sheet_names)} sheets)...")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        try:
            price_col_index = headers.index(price_col_name)
        except ValueError:
            print(f"'{price_col_name}' not found in sheet: {sheet_name}")
            continue

        merged_cells_to_replace = set()
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col - 1 == price_col_index:
                for row in range(merged_range.min_row + 1, merged_range.max_row + 1):
                    merged_cells_to_replace.add((row, merged_range.min_col))

        data = []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = []
            for j, cell in enumerate(row):
                if j == price_col_index and (i, j + 1) in merged_cells_to_replace:
                    row_data.append("NO_PRICE")
                else:
                    row_data.append(cell.value)
            data.append(row_data)

        df = pd.DataFrame(data, columns=headers)

        df.iloc[:, 0] = df.iloc[:, 0].replace(r'^\s*$', np.nan, regex=True).ffill()
        df.iloc[:, 1] = df.iloc[:, 1].replace(r'^\s*$', np.nan, regex=True).ffill()

        if col_name_info in df.columns:
            df[col_name_info] = df[col_name_info].replace(r'^\s*$', np.nan, regex=True).ffill()

        safe_sheet_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in sheet_name)
        csv_file = os.path.join(output_folder, f"{safe_sheet_name}.csv")

        df.to_csv(csv_file, index=False, encoding="utf-8")
        print(f"Saved: {csv_file}")

    print(f"Conversion complete for {city.UPPER()}!")

def main():
    for city, xlsx_path in original_files.items():
        process_city(city, xlsx_path)
    print("\n All conversions complete!")

if __name__ == "__main__":
    main()
